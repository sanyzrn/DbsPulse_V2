"""P1-08 — میانگینی که روی جمعیت کوچک حساب شده، آمار نیست؛ افشای فرد است.

اگر یک واحد سازمانی دو نفر داشته باشد، «میانگین واحد» تقریباً امتیاز همان دو نفر
است؛ اگر یک نفر داشته باشد، *دقیقاً* امتیاز اوست، فقط با برچسبی که بی‌نام به‌نظر
می‌رسد. تا وقتی گزارش‌ها فقط برای HR است این بیشتر محافظت از فایل‌هایی است که از دست
HR خارج می‌شوند، ولی پیش‌نیاز سختِ باز کردن آنالیتیکس به نقش‌های دیگر است.
"""
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.core.config import settings
from app.services.privacy import is_below_cohort, suppressed_avg
from tests.helpers import (
    active_indicators,
    auth_header,
    full_valid_scores,
    make_access,
    make_personnel,
    make_user,
)


def _finalized_evaluation(client, db_session, org_unit: str, hr, dep, ceo):
    sup = make_user(db_session, "unit_supervisor")
    personnel = make_personnel(db_session, org_unit=org_unit)
    make_access(db_session, personnel, sup, dep, ceo)
    db_session.commit()

    evaluation = client.post(
        "/api/evaluations", json={"subject_personnel_id": personnel.id}, headers=auth_header(sup)
    ).json()
    client.put(
        f"/api/evaluations/{evaluation['id']}/scores",
        json={"scores": full_valid_scores(active_indicators(db_session))},
        headers=auth_header(sup),
    )
    client.post(f"/api/evaluations/{evaluation['id']}/submit", headers=auth_header(sup))
    client.post(f"/api/evaluations/{evaluation['id']}/hr-approve", headers=auth_header(hr))
    client.post(f"/api/evaluations/{evaluation['id']}/deputy-approve", headers=auth_header(dep))
    client.post(f"/api/evaluations/{evaluation['id']}/ceo-finalize", headers=auth_header(ceo))
    return personnel


@pytest.fixture()
def small_unit(client, db_session):
    """یک واحد سازمانی با تعداد ارزیابیِ کمتر از آستانه."""
    hr = make_user(db_session, "hr")
    dep = make_user(db_session, "deputy")
    ceo = make_user(db_session, "ceo")
    db_session.commit()

    unit = "واحد کوچک تست"
    personnel = _finalized_evaluation(client, db_session, unit, hr, dep, ceo)
    return {"hr": hr, "unit": unit, "personnel": personnel}


# ------------------------------------------------------------ pure helper


def test_helper_suppresses_below_the_threshold():
    below = settings.min_cohort_size - 1
    assert suppressed_avg(82.5, below) is None
    assert is_below_cohort(below) is True


def test_helper_passes_values_at_or_above_the_threshold():
    at = settings.min_cohort_size
    assert suppressed_avg(82.5, at) == 82.5
    assert is_below_cohort(at) is False


def test_helper_keeps_none_as_none():
    assert suppressed_avg(None, 1000) is None


# ---------------------------------------------------------------- reports


def test_small_unit_average_is_suppressed_in_the_summary(client, db_session, small_unit):
    r = client.get(
        "/api/dashboard/report/summary",
        params={"org_unit": small_unit["unit"]},
        headers=auth_header(small_unit["hr"]),
    )

    assert r.status_code == 200
    body = r.json()
    row = next(u for u in body["by_org_unit"] if u["org_unit"] == small_unit["unit"])
    assert row["avg_final_pct"] is None, "میانگینِ یک واحد تک‌نفره نباید نمایش داده شود"
    # تعداد پنهان نمی‌شود: دانستن «این واحد یک ارزیابی دارد» افشای عملکرد نیست
    assert row["count"] >= 1


def test_the_overall_average_is_suppressed_for_a_small_filtered_set(
    client, db_session, small_unit
):
    r = client.get(
        "/api/dashboard/report/summary",
        params={"org_unit": small_unit["unit"]},
        headers=auth_header(small_unit["hr"]),
    )

    assert r.json()["avg_final_pct"] is None


def test_naming_a_person_explicitly_is_not_suppressed(client, db_session, small_unit):
    """اگر کاربر خودش نام فرد را داده، چیزی «کشف» نمی‌شود — سرکوب فقط اذیت می‌کرد."""
    r = client.get(
        "/api/dashboard/report/summary",
        params={"personnel_id": small_unit["personnel"].id},
        headers=auth_header(small_unit["hr"]),
    )

    assert r.json()["avg_final_pct"] is not None


def test_indicator_breakdown_suppresses_small_units(client, db_session, small_unit):
    indicator = active_indicators(db_session)[0]

    r = client.get(
        f"/api/dashboard/report/indicator/{indicator.id}",
        params={"org_unit": small_unit["unit"]},
        headers=auth_header(small_unit["hr"]),
    )

    assert r.status_code == 200
    for row in r.json()["by_org_unit"]:
        if row["count"] < settings.min_cohort_size:
            assert row["avg_score"] is None


def test_employee_vs_unit_hides_the_unit_average_but_keeps_the_persons_own_scores(
    client, db_session, small_unit
):
    r = client.get(
        "/api/dashboard/report/employee-vs-unit",
        params={"personnel_id": small_unit["personnel"].id},
        headers=auth_header(small_unit["hr"]),
    )

    body = r.json()
    assert body["unit_avg"] is None, "میانگین یک واحد تک‌نفره همان امتیاز آن یک نفر است"
    assert body["employee_avg"] is not None, "امتیاز خودِ فرد به کاربری که نامش را داده نمایش داده می‌شود"
    assert body["per_evaluation"], "سری امتیازهای خودِ فرد نباید سرکوب شود"


# ----------------------------------------------------------------- export


def test_the_excel_export_marks_suppressed_cells_instead_of_leaking_or_crashing(
    client, db_session, small_unit
):
    r = client.get(
        "/api/dashboard/report/export.xlsx",
        params={"org_unit": small_unit["unit"]},
        headers=auth_header(small_unit["hr"]),
    )
    assert r.status_code == 200

    sheet = load_workbook(BytesIO(r.content))["میانگین به‌تفکیک واحد"]
    row = next(
        r for r in sheet.iter_rows(min_row=2, values_only=True) if r[0] == small_unit["unit"]
    )
    assert isinstance(row[1], str) and "کمتر از حد نمایش" in row[1]


# -------------------------------------------------------------- dashboard


def test_dashboard_overview_suppresses_small_units_and_ranks_only_visible_ones(
    client, db_session, small_unit
):
    r = client.get("/api/dashboard/overview", headers=auth_header(small_unit["hr"]))

    assert r.status_code == 200
    body = r.json()
    small = next(u for u in body["by_org_unit"] if u["org_unit"] == small_unit["unit"])
    assert small["avg_final_pct"] is None
    # فهرست «ضعیف‌ترین واحدها» نباید واحد سرکوب‌شده را رتبه‌بندی کند — خودِ حضور در
    # آن فهرست یک نشت است.
    assert all(u["avg_final_pct"] is not None for u in body["lowest_by_unit"])


def test_turning_the_threshold_down_restores_the_numbers(
    client, db_session, small_unit, no_cohort_suppression
):
    """آستانه پیکربندی‌پذیر است تا محیط‌های کوچک بتوانند آگاهانه خاموشش کنند."""
    r = client.get(
        "/api/dashboard/report/summary",
        params={"org_unit": small_unit["unit"]},
        headers=auth_header(small_unit["hr"]),
    )

    row = next(u for u in r.json()["by_org_unit"] if u["org_unit"] == small_unit["unit"])
    assert row["avg_final_pct"] is not None
