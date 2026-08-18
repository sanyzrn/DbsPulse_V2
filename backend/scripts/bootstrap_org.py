"""حساب کاربریِ ارزیاب‌ها را از روی همان فایل پرسنلی می‌سازد.

چرا این فایل هست
----------------
مسیر راه‌اندازی واقعی سه قدم دارد و قدم دوم تنها جایی بود که ابزار نداشت:

    ۱. scripts.create_admin      → اولین حساب مدیر سامانه
    ۲. (این‌جا)                   → حساب ارزیاب‌ها
    ۳. بارگذاری اکسل از پنل      → پرسنل و زنجیرهٔ ارزیابی

بدون قدم دوم، بارگذاری فایل هر ردیف را با خطای «کاربری با نام … پیدا نشد» رد
می‌کند — چون زنجیره به *حساب* اشاره می‌کند، نه به متن.

**چرا SQL نمی‌دهد.** رمزها در دیتابیس نیستند؛ فقط هشِ Argon2id ذخیره می‌شود که
یک‌طرفه است. پس فایل SQL آماده‌ای که رمز داشته باشد اصلاً وجود ندارد — رمز باید
همین‌جا ساخته و هش شود. علاوه بر آن، INSERT دستی از گاردهای برنامه (نقشِ مجاز
برای هر مرحله، تداخل ارزیاب و ارزیابی‌شونده) رد نمی‌شود.

دو مرحله‌ای است، و عمداً
------------------------
بار اول فقط *نقشه* را می‌نویسد: چه کسانی پیدا شدند، هرکدام چه نقشی می‌گیرند و
چه نام کاربری‌ای پیشنهاد می‌شود. نام کاربری چیزی است که به آدم واقعی تحویل
می‌دهید؛ حدسِ ماشین از روی نام فارسی نباید بی‌آنکه ببینیدش قطعی شود.

    python -m scripts.bootstrap_org --file personnel.xlsx --plan evaluators.csv
    # ستون username را در فایل اصلاح کنید، بعد:
    python -m scripts.bootstrap_org --apply evaluators.csv

رمزها یک‌بار روی صفحه چاپ می‌شوند و هیچ‌جا ذخیره نمی‌شوند. هر حساب با اجبار به
تغییر رمز در اولین ورود ساخته می‌شود.
"""
from __future__ import annotations

import argparse
import csv
import secrets
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.security import hash_password
from app.db.session import SessionLocal
from app.models.enums import UserRole
from app.models.user import User

#: ستون فایل → نقشی که آن ستون می‌سازد، از پایین به بالا. ترتیب مهم است: کسی که
#: در چند ستون آمده، *بالاترین* نقش را می‌گیرد. قانون سلسله‌مراتب اجازه می‌دهد
#: مافوق کارِ مرحلهٔ پایین‌تر را بکند، پس بالاترین نقش همهٔ جایگاه‌هایش را پوشش
#: می‌دهد؛ برعکسش نه.
_COLUMN_ROLES: tuple[tuple[str, UserRole], ...] = (
    ("مسئول مستقیم", UserRole.unit_supervisor),
    ("معاونت مربوطه", UserRole.deputy),
    ("مدیرعامل", UserRole.ceo),
)

_RANK = {UserRole.unit_supervisor: 1, UserRole.deputy: 2, UserRole.ceo: 3}

#: مقادیری که یعنی «این مرحله را ندارد».
_ABSENT = {"", "-", "—", "ندارد", "نامشخص"}

#: نویسه‌گردانی سادهٔ فارسی→لاتین، فقط برای *پیشنهاد* نام کاربری. کامل نیست و
#: قرار هم نیست باشد: خروجی‌اش را آدم می‌بیند و اصلاح می‌کند.
_TRANSLITERATE = {
    "ا": "a", "آ": "a", "ب": "b", "پ": "p", "ت": "t", "ث": "s", "ج": "j",
    "چ": "ch", "ح": "h", "خ": "kh", "د": "d", "ذ": "z", "ر": "r", "ز": "z",
    "ژ": "zh", "س": "s", "ش": "sh", "ص": "s", "ض": "z", "ط": "t", "ظ": "z",
    "ع": "a", "غ": "gh", "ف": "f", "ق": "gh", "ک": "k", "گ": "g", "ل": "l",
    "م": "m", "ن": "n", "و": "v", "ه": "h", "ی": "y", "ئ": "y", "ة": "h",
}


def _suggest_username(full_name: str) -> str:
    """آخرین بخش نام (نام خانوادگی) به لاتین — کوتاه و قابل تایپ."""
    parts = [p for p in full_name.replace("‌", " ").split() if p]
    source = parts[-1] if parts else full_name
    latin = "".join(_TRANSLITERATE.get(ch, "") for ch in source)
    latin = unicodedata.normalize("NFKD", latin).encode("ascii", "ignore").decode()
    return (latin or "user")[:20].lower()


def collect_evaluators(path: Path) -> dict[str, UserRole]:
    """نام هر ارزیاب، با بالاترین نقشی که در فایل برایش لازم است."""
    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        sys.exit("فایل خالی است.")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]

    found: dict[str, UserRole] = {}
    for row in rows[1:]:
        for column, role in _COLUMN_ROLES:
            if column not in header:
                continue
            position = header.index(column)
            if position >= len(row):
                continue
            name = str(row[position]).strip() if row[position] is not None else ""
            if name in _ABSENT:
                continue
            if name not in found or _RANK[role] > _RANK[found[name]]:
                found[name] = role
    return found


def write_plan(path: Path, evaluators: dict[str, UserRole]) -> None:
    used: set[str] = set()
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["full_name", "role", "username"])
        for name, role in sorted(evaluators.items(), key=lambda i: -_RANK[i[1]]):
            suggestion = _suggest_username(name)
            candidate, n = suggestion, 2
            while candidate in used:
                candidate, n = f"{suggestion}{n}", n + 1
            used.add(candidate)
            writer.writerow([name, role.value, candidate])
    print(f"نقشه در «{path}» نوشته شد: {len(evaluators)} ارزیاب")
    print("ستون username را بررسی/اصلاح کنید، بعد با --apply همان فایل را بدهید.")


def apply_plan(path: Path) -> int:
    with path.open(encoding="utf-8-sig") as handle:
        plan = list(csv.DictReader(handle))
    if not plan:
        sys.exit("فایل نقشه خالی است.")

    created: list[tuple[str, str, str]] = []
    with SessionLocal() as db:
        for entry in plan:
            full_name = (entry.get("full_name") or "").strip()
            username = (entry.get("username") or "").strip()
            role_value = (entry.get("role") or "").strip()
            if not (full_name and username and role_value):
                print(f"    رد شد (ستون ناقص): {entry}")
                continue
            if db.scalar(select(User).where(User.username == username)) is not None:
                print(f"    «{username}» از قبل هست — دست‌نخورده ماند")
                continue

            password = secrets.token_urlsafe(9)
            db.add(
                User(
                    username=username,
                    full_name=full_name,
                    password_hash=hash_password(password),
                    role=UserRole(role_value),
                    is_active=True,
                    # رمز را ما ساخته‌ایم و روی صفحه چاپ می‌شود؛ صاحب حساب باید
                    # در اولین ورود عوضش کند.
                    must_change_password=True,
                )
            )
            created.append((username, full_name, password))
        db.commit()

    if not created:
        print("هیچ حساب تازه‌ای ساخته نشد.")
        return 0
    print("\nاین رمزها فقط همین یک‌بار نشان داده می‌شوند:\n")
    print(f"{'نام کاربری':<20} {'رمز موقت':<16} نام")
    for username, full_name, password in created:
        print(f"{username:<20} {password:<16} {full_name}")
    print(f"\n{len(created)} حساب ساخته شد. حالا فایل پرسنلی را از پنل بارگذاری کنید.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="ساخت حساب ارزیاب‌ها از روی فایل پرسنلی")
    parser.add_argument("--file", type=Path, help="فایل اکسل پرسنل")
    parser.add_argument("--plan", type=Path, help="مسیر نوشتن نقشه (CSV)")
    parser.add_argument("--apply", type=Path, help="ساخت حساب‌ها از روی نقشهٔ اصلاح‌شده")
    args = parser.parse_args()

    if args.apply:
        return apply_plan(args.apply)
    if not (args.file and args.plan):
        parser.error("یا --file و --plan بدهید، یا --apply")
    evaluators = collect_evaluators(args.file)
    if not evaluators:
        sys.exit("هیچ نامی در ستون‌های زنجیره پیدا نشد.")
    write_plan(args.plan, evaluators)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
