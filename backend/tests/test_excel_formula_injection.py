"""P2-02 — متن آزادِ کاربر نباید در فایل اکسل به فرمول تبدیل شود.

سناریو: یک مسئول واحد در فیلد «شواهد» یا عنوان برنامهٔ بهبود چیزی مثل
`=cmd|'/c calc'!A1` می‌نویسد. HR فایل اکسل را باز می‌کند و Excel آن را اجرا می‌کند
(یا در حالت خفیف‌تر، openpyxl همان موقع سلول را به‌عنوان فرمول می‌نویسد).
"""
from datetime import date, datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.excel import _neutralise, build_personnel_workbook, build_report_workbook


class _FakePersonnel:
    def __init__(self, full_name: str, job_title: str = "کارشناس"):
        self.personnel_code = "P-1"
        self.full_name = full_name
        self.job_title = job_title
        self.org_unit = "واحد"
        self.is_manager = False
        self.status = type("S", (), {"value": "active"})()
        self.contract_start_date = date(2025, 1, 1)
        self.contract_end_date = date(2026, 1, 1)
        self.created_at = datetime(2025, 1, 1)


@pytest.mark.parametrize(
    "dangerous",
    [
        "=cmd|'/c calc'!A1",
        "+1+1",
        "-2+3",
        "@SUM(A1)",
        "\t=1+1",
    ],
)
def test_formula_triggering_prefixes_are_neutralised(dangerous):
    assert _neutralise(dangerous) == "'" + dangerous


@pytest.mark.parametrize("harmless", ["علی محمدی", "واحد فروش", "", "۱۲۳", "A=B"])
def test_ordinary_text_is_left_alone(harmless):
    assert _neutralise(harmless) == harmless


@pytest.mark.parametrize("number", [0, 5, -3, 4.25, None, True])
def test_non_string_values_are_left_alone(number):
    assert _neutralise(number) is number


def test_exported_cell_is_text_not_a_formula():
    """مهم‌ترین ادعا: سلول در فایل واقعی از نوع فرمول نیست."""
    payload = "=cmd|'/c calc'!A1"
    workbook = load_workbook(
        BytesIO(build_personnel_workbook([_FakePersonnel(payload)]))
    )
    cell = workbook.active.cell(row=2, column=2)

    assert cell.data_type != "f", "سلول به‌عنوان فرمول نوشته شده است"
    assert cell.value == "'" + payload


def test_secondary_sheets_of_the_report_are_guarded_too():
    """برگه‌های دوم/سوم گزارش هم از همان مسیر امن ساخته می‌شوند."""
    workbook = load_workbook(
        BytesIO(
            build_report_workbook(
                total=1,
                avg_final_pct=80.0,
                by_org_unit=[("=HYPERLINK(\"http://evil\")", 80.0, 1)],
                by_indicator=[("@general", "-درصد رشد", 4.0, 1)],
            )
        )
    )

    unit_cell = workbook["میانگین به‌تفکیک واحد"].cell(row=2, column=1)
    assert unit_cell.data_type != "f"
    assert unit_cell.value.startswith("'=HYPERLINK")

    indicator_sheet = workbook["میانگین به‌تفکیک شاخص"]
    assert indicator_sheet.cell(row=2, column=1).value == "'@general"
    assert indicator_sheet.cell(row=2, column=2).value == "'-درصد رشد"


def test_numbers_in_the_report_stay_numeric():
    """ضدعفونی نباید ستون‌های عددی را به متن تبدیل کند، وگرنه نمودار/میانگین می‌شکند."""
    workbook = load_workbook(
        BytesIO(
            build_report_workbook(
                total=3,
                avg_final_pct=72.5,
                by_org_unit=[("فروش", 72.5, 3)],
                by_indicator=[],
            )
        )
    )

    unit_sheet = workbook["میانگین به‌تفکیک واحد"]
    assert unit_sheet.cell(row=2, column=2).value == 72.5
    assert unit_sheet.cell(row=2, column=3).value == 3


def test_a_payload_imported_from_excel_comes_back_out_neutralised(client, db_session):
    """ورود دسته‌ای، یک مسیر ورودیِ *تازه* برای همین حمله است.

    P2-02 وقتی نوشته شد که تنها راه ورود متن آزاد، فرم‌های خود برنامه بود. حالا
    HR می‌تواند فایل اکسل بارگذاری کند — و فایل اکسل دقیقاً همان جایی است که این
    payloadها زندگی می‌کنند. این تست کل زنجیره را می‌سنجد: ورود → ذخیره → خروجی.
    """
    from openpyxl import Workbook

    from app.services.personnel_import import COLUMNS
    from tests.helpers import auth_header, make_user

    hr = make_user(db_session, "hr")
    db_session.commit()

    payload = "=cmd|'/c calc'!A1"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(COLUMNS)
    sheet.append(
        ["P-INJ-1", "جای‌نگه‌دار", "کارشناس", "واحد تست", "خیر", "فعال", "۱۴۰۵/۰۱/۰۱", "۱۴۰۶/۰۱/۰۱", ""]
    )
    # نام را به‌صورت *متن* می‌نویسیم، نه فرمول. مهاجم واقعی هم همین کار را می‌کند:
    # اگر سلول واقعاً فرمول باشد، خواندن با data_only=True مقدار کش‌شده را می‌خواهد
    # که در فایلِ هرگز-بازنشده None است و ردیف همان‌جا رد می‌شود (تست بعدی). خطر
    # واقعی متنی است که *بعداً* هنگام خروجی‌گرفتن به فرمول تبدیل شود.
    name_cell = sheet.cell(row=2, column=2)
    name_cell.value = payload
    name_cell.data_type = "s"
    buffer = BytesIO()
    workbook.save(buffer)

    imported = client.post(
        "/api/personnel/import",
        files={
            "file": (
                "attack.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(hr),
    )
    assert imported.status_code == 200
    assert imported.json()["created_personnel"] == 1

    exported = client.get("/api/personnel/export.xlsx", headers=auth_header(hr))
    assert exported.status_code == 200
    out = load_workbook(BytesIO(exported.content)).active
    cells = [value for row in out.iter_rows(values_only=True) for value in row]

    # مقدار خام هرگز نباید در فایل باشد؛ فقط نسخهٔ خنثی‌شده با آپاستروف
    assert payload not in cells
    assert "'" + payload in cells


def test_a_real_formula_cell_is_rejected_rather_than_imported_as_something_odd(
    client, db_session
):
    """اگر سلول *واقعاً* فرمول باشد، خواندن با data_only مقدار کش‌شده را می‌خواهد که
    در فایلِ هرگز-بازنشده وجود ندارد. نتیجه باید یک ردِ روشن باشد، نه ردیفی با نام
    خالی یا مقداری که کاربر ننوشته."""
    from openpyxl import Workbook

    from app.services.personnel_import import COLUMNS
    from tests.helpers import auth_header, make_user

    hr = make_user(db_session, "hr")
    db_session.commit()

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(COLUMNS)
    sheet.append(
        ["P-INJ-2", "=SUM(A1:A9)", "کارشناس", "واحد", "خیر", "فعال", "۱۴۰۵/۰۱/۰۱", "۱۴۰۶/۰۱/۰۱", ""]
    )
    buffer = BytesIO()
    workbook.save(buffer)

    preview = client.post(
        "/api/personnel/import/preview",
        files={
            "file": (
                "formula.xlsx",
                buffer.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(hr),
    ).json()

    assert preview["valid_count"] == 0
    assert any("نام و نام خانوادگی" in error for error in preview["rows"][0]["errors"])
