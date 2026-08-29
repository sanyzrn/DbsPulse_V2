"""ورود دسته‌ای پرسنل از Excel.

راه‌اندازی یک مشتری تازه یعنی وارد کردن ده‌ها یا صدها پرسنل؛ تایپ دستی هم کند
است هم مستعد خطا، و غلط تایپی در «کد پرسنلی» بعداً به پروندهٔ ارزیابیِ گره‌خورده
به فرد اشتباه تبدیل می‌شود.

دو خاصیتِ مهم که این‌جا سنجیده می‌شوند:

* پیش‌نمایش **هیچ چیزی نمی‌نویسد** — وگرنه «ببین چه می‌شود» خودش همان اتفاق است.
* خروجی این سامانه باید دوباره ورودی همین سامانه باشد (رفت‌وبرگشت اکسل)، چون
  جریان واقعی HR همین است: خروجی بگیر، در اکسل ویرایش کن، برگردان.
"""
from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models.enums import Capability
from app.models.personnel import Personnel
from app.models.user import User
from app.services.personnel_import import COLUMNS, parse_flexible_date
from tests.helpers import auth_header, make_personnel, make_user

HEADER = COLUMNS


def _sheet(rows: list[list], header: list[str] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header if header is not None else HEADER)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _row(code="P-9001", name="کارمند وارداتی", username="", site=""):
    # ترتیب = ترتیبِ COLUMNS. «محل» ستون تازه‌ای است بین «عنوان شغلی» و «واحد
    # سازمانی» و اختیاری است، پس این‌جا خالی می‌ماند مگر تست خودش پرش کند.
    return [code, name, "کارشناس", site, "واحد تست", "خیر", "فعال", "۱۴۰۵/۰۱/۰۱", "۱۴۰۶/۰۱/۰۱", username]


def _upload(client, hr, path, content, filename="import.xlsx"):
    return client.post(
        path,
        files={
            "file": (
                filename,
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        headers=auth_header(hr),
    )


# ───────────────────────────────── تجزیهٔ تاریخ


def test_jalali_and_gregorian_and_persian_digits_all_parse():
    assert parse_flexible_date("۱۴۰۵/۰۱/۰۱") == date(2026, 3, 21)
    assert parse_flexible_date("1405/01/01") == date(2026, 3, 21)
    assert parse_flexible_date("2026-03-21") == date(2026, 3, 21)
    assert parse_flexible_date(date(2026, 3, 21)) == date(2026, 3, 21)


def test_unparseable_dates_return_none_instead_of_guessing():
    assert parse_flexible_date("دیروز") is None
    assert parse_flexible_date("1405/13/01") is None  # ماه ۱۳ وجود ندارد
    assert parse_flexible_date("") is None


# ───────────────────────────────── پیش‌نمایش


def test_preview_reports_what_would_happen_and_writes_nothing(client, db_session):
    hr = make_user(db_session, "hr")
    db_session.commit()
    before = db_session.scalar(select(Personnel).where(Personnel.personnel_code == "P-9001"))
    assert before is None

    r = _upload(client, hr, "/api/personnel/import/preview", _sheet([_row()]))

    assert r.status_code == 200
    body = r.json()
    assert body["valid_count"] == 1
    assert body["invalid_count"] == 0
    # و مهم‌تر: هنوز چیزی ساخته نشده
    assert db_session.scalar(select(Personnel).where(Personnel.personnel_code == "P-9001")) is None


def test_preview_names_the_row_and_the_reason(client, db_session):
    hr = make_user(db_session, "hr")
    db_session.commit()

    content = _sheet(
        [
            _row(code="P-9001"),
            ["", "بدون کد", "کارشناس", "واحد", "خیر", "فعال", "۱۴۰۵/۰۱/۰۱", "۱۴۰۶/۰۱/۰۱", ""],
        ]
    )
    body = _upload(client, hr, "/api/personnel/import/preview", content).json()

    assert body["valid_count"] == 1
    assert body["invalid_count"] == 1
    bad = next(r for r in body["rows"] if r["errors"])
    # ردیف ۳ فایل: سرستون ۱، ردیف سالم ۲
    assert bad["row_number"] == 3
    assert any("کد پرسنلی" in e for e in bad["errors"])


def test_duplicate_code_inside_the_file_is_caught(client, db_session):
    hr = make_user(db_session, "hr")
    db_session.commit()

    body = _upload(
        client, hr, "/api/personnel/import/preview", _sheet([_row(), _row()])
    ).json()

    assert body["valid_count"] == 1
    assert body["invalid_count"] == 1
    assert any("تکراری" in e for e in body["rows"][1]["errors"])


def test_code_that_already_exists_is_caught(client, db_session):
    hr = make_user(db_session, "hr")
    existing = make_personnel(db_session, personnel_code="P-9001")
    db_session.commit()
    assert existing.personnel_code == "P-9001"

    body = _upload(client, hr, "/api/personnel/import/preview", _sheet([_row()])).json()

    assert body["invalid_count"] == 1
    assert any("از قبل" in e for e in body["rows"][0]["errors"])


def test_contract_end_before_start_is_rejected(client, db_session):
    hr = make_user(db_session, "hr")
    db_session.commit()
    row = _row()
    row[7], row[8] = "۱۴۰۶/۰۱/۰۱", "۱۴۰۵/۰۱/۰۱"

    body = _upload(client, hr, "/api/personnel/import/preview", _sheet([row])).json()

    assert any("بعد از" in e for e in body["rows"][0]["errors"])


def test_missing_columns_are_reported_as_a_file_error(client, db_session):
    hr = make_user(db_session, "hr")
    db_session.commit()

    body = _upload(
        client, hr, "/api/personnel/import/preview", _sheet([["P-1"]], header=["کد پرسنلی"])
    ).json()

    assert body["file_errors"]
    assert "نام و نام خانوادگی" in body["file_errors"][0]


def test_a_non_excel_file_is_refused_by_extension(client, db_session):
    hr = make_user(db_session, "hr")
    db_session.commit()

    r = _upload(client, hr, "/api/personnel/import/preview", b"not a workbook", filename="x.csv")

    assert r.status_code == 400


def test_a_corrupt_xlsx_gives_a_readable_message_not_a_500(client, db_session):
    hr = make_user(db_session, "hr")
    db_session.commit()

    r = _upload(client, hr, "/api/personnel/import/preview", b"definitely not a zip")

    assert r.status_code == 200
    assert r.json()["file_errors"]


def test_blank_trailing_rows_are_ignored(client, db_session):
    """اکسل معمولاً چند سطر خالی ته فایل دارد؛ اینها نه داده‌اند نه خطا."""
    hr = make_user(db_session, "hr")
    db_session.commit()

    body = _upload(
        client, hr, "/api/personnel/import/preview", _sheet([_row(), [None] * 10, ["", "", ""]])
    ).json()

    assert body["total_rows"] == 1
    assert body["valid_count"] == 1


# ───────────────────────────────── درج


def test_commit_creates_personnel_and_accounts(client, db_session):
    hr = make_user(db_session, "hr")
    db_session.commit()

    content = _sheet(
        [
            _row(code="P-9001", name="بدون حساب"),
            _row(code="P-9002", name="با حساب", username="karbar.jadid"),
        ]
    )
    r = _upload(client, hr, "/api/personnel/import", content)

    assert r.status_code == 200
    body = r.json()
    assert body["created_personnel"] == 2
    assert body["created_accounts"] == 1
    assert body["skipped_rows"] == 0

    created = db_session.scalar(select(Personnel).where(Personnel.personnel_code == "P-9002"))
    assert created is not None and created.full_name == "با حساب"
    user = db_session.scalar(select(User).where(User.username == "karbar.jadid"))
    assert user is not None
    assert user.personnel_id == created.id
    # رمز موقت است: کاربر در اولین ورود مجبور به تغییرش می‌شود
    assert user.must_change_password is True


def test_the_temporary_password_comes_back_once_and_actually_works(client, db_session):
    hr = make_user(db_session, "hr")
    db_session.commit()

    body = _upload(
        client,
        hr,
        "/api/personnel/import",
        _sheet([_row(code="P-9003", username="karbar.login")]),
    ).json()

    account = body["accounts"][0]
    assert account["username"] == "karbar.login"
    assert len(account["temporary_password"]) >= 10

    # رمز برگشتی واقعاً کار می‌کند، وگرنه HR چیزی را توزیع می‌کند که کار نمی‌کند
    login = client.post(
        "/api/auth/login",
        json={"username": "karbar.login", "password": account["temporary_password"]},
    )
    assert login.status_code == 200
    assert login.json()["must_change_password"] is True


def test_invalid_rows_are_skipped_and_the_good_ones_still_land(client, db_session):
    """یک غلط تایپی در ردیف آخر نباید ردیف‌های درستِ قبلی را دور بریزد."""
    hr = make_user(db_session, "hr")
    db_session.commit()

    content = _sheet(
        [
            _row(code="P-9101"),
            _row(code="P-9102"),
            ["", "بی‌کد", "کارشناس", "واحد", "خیر", "فعال", "۱۴۰۵/۰۱/۰۱", "۱۴۰۶/۰۱/۰۱", ""],
        ]
    )
    body = _upload(client, hr, "/api/personnel/import", content).json()

    assert body["created_personnel"] == 2
    assert body["skipped_rows"] == 1
    assert db_session.scalar(select(Personnel).where(Personnel.personnel_code == "P-9101"))
    assert db_session.scalar(select(Personnel).where(Personnel.personnel_code == "P-9102"))


def test_the_password_never_reaches_the_audit_log(client, db_session):
    """لاگ ممیزی ماندگار و append-only است؛ رمز نباید ماندگار شود."""
    hr = make_user(
        db_session,
        "hr",
        capabilities=[Capability.manage_personnel, Capability.view_audit_log],
    )
    db_session.commit()

    body = _upload(
        client,
        hr,
        "/api/personnel/import",
        _sheet([_row(code="P-9200", username="karbar.secret")]),
    ).json()
    password = body["accounts"][0]["temporary_password"]

    entries = client.get(
        "/api/audit-log", params={"limit": 100}, headers=auth_header(hr)
    ).json()["items"]
    assert entries
    assert password not in str(entries)


def test_import_is_logged_with_the_counts(client, db_session):
    hr = make_user(
        db_session,
        "hr",
        capabilities=[Capability.manage_personnel, Capability.view_audit_log],
    )
    db_session.commit()

    _upload(client, hr, "/api/personnel/import", _sheet([_row(code="P-9300")]))

    rows = client.get(
        "/api/audit-log", params={"event_type": "personnel_imported"}, headers=auth_header(hr)
    ).json()["items"]
    assert rows
    assert rows[0]["new_value"]["created_personnel"] == 1


# ───────────────────────────────── رفت‌وبرگشت و دسترسی


def test_the_export_can_be_imported_again(client, db_session):
    """جریان واقعی HR: خروجی بگیر، در اکسل ویرایش کن، برگردان.

    اگر ورودی همان چیزی را که خروجی تولید می‌کند نپذیرد — ارقام فارسی، تاریخ
    شمسی، «بله/خیر» — این چرخه همان‌جا می‌شکند.
    """
    hr = make_user(db_session, "hr")
    make_personnel(db_session, personnel_code="P-8001", full_name="نفر موجود")
    db_session.commit()

    exported = client.get("/api/personnel/export.xlsx", headers=auth_header(hr))
    assert exported.status_code == 200

    # کدها را عوض می‌کنیم تا «تکراری» نباشند؛ باقی ستون‌ها دقیقاً همان خروجی‌اند
    workbook = load_workbook(BytesIO(exported.content))
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_col=1):
        if row[0].value:
            row[0].value = f"{row[0].value}-COPY"
    buffer = BytesIO()
    workbook.save(buffer)

    body = _upload(client, hr, "/api/personnel/import/preview", buffer.getvalue()).json()

    assert body["file_errors"] == []
    assert body["invalid_count"] == 0, body["rows"]
    assert body["valid_count"] >= 1


def test_the_template_is_itself_importable(client, db_session):
    """فایل نمونه باید بدون ویرایش هم از اعتبارسنجی رد شود، وگرنه راهنما نیست."""
    hr = make_user(db_session, "hr")
    db_session.commit()

    template = client.get("/api/personnel/import-template.xlsx", headers=auth_header(hr))
    assert template.status_code == 200

    body = _upload(client, hr, "/api/personnel/import/preview", template.content).json()

    assert body["file_errors"] == []
    assert body["valid_count"] == 1


def test_only_hr_may_import(client, db_session):
    supervisor = make_user(db_session, "unit_supervisor")
    db_session.commit()

    for path in ("/api/personnel/import/preview", "/api/personnel/import"):
        assert _upload(client, supervisor, path, _sheet([_row()])).status_code == 403
    assert (
        client.get("/api/personnel/import-template.xlsx", headers=auth_header(supervisor)).status_code
        == 403
    )
