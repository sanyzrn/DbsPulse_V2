"""ورود دسته‌ای پرسنل از فایل Excel — تجزیه و اعتبارسنجی، بدون هیچ نوشتنی.

راه‌اندازی یک مشتری تازه یعنی وارد کردن ده‌ها یا صدها پرسنل. تایپ دستی هم کند
است هم مستعد خطا، و خطای تایپی در «کد پرسنلی» بعداً به یک پروندهٔ ارزیابی
گره‌خورده به فرد اشتباه تبدیل می‌شود.

دو تصمیم که شکل این ماژول را تعیین کرده‌اند:

۱. **اعتبارسنجی کاملاً از درج جداست.** این فایل هیچ‌وقت چیزی نمی‌نویسد. خروجی‌اش
   یک گزارش است: کدام ردیف‌ها سالم‌اند، کدام‌ها نیستند و دقیقاً چرا. UI همین را
   نشان می‌دهد و تا تأیید کاربر چیزی درج نمی‌شود — چون «۲۰۰ ردیف وارد شد و ۳تا
   اشتباه بود» را نمی‌شود به‌سادگی برگرداند.

۲. **ورودی باید همان چیزی را بپذیرد که خروجی تولید می‌کند.** جریان واقعی HR این
   است: خروجی اکسل بگیر، در اکسل ویرایش کن، دوباره وارد کن. پس ارقام فارسی،
   تاریخ شمسی و «بله/خیر» و «فعال/غیرفعال» همگی پذیرفته می‌شوند. یک تست همین
   رفت‌وبرگشت را می‌سنجد.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO

import jdatetime
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.enums import PersonnelStatus
from app.models.personnel import Personnel
from app.models.user import User

# همان ستون‌های build_personnel_workbook، به‌علاوهٔ یک ستون اختیاری برای نام کاربری
COLUMNS = [
    "کد پرسنلی",
    "نام و نام خانوادگی",
    "عنوان شغلی",
    "واحد سازمانی",
    "مدیر",
    "وضعیت",
    "شروع قرارداد",
    "پایان قرارداد",
    "نام کاربری",
]
REQUIRED_COLUMNS = COLUMNS[:4] + COLUMNS[6:8]

# ارقام فارسی و عربی → اسکی. بدون این، «۱۴۰۵/۰۱/۰۱» که خودِ ما تولید کرده‌ایم
# هنگام بازگشت غیرقابل تجزیه می‌شد.
_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")

_TRUE_WORDS = {"بله", "بلی", "آری", "true", "yes", "1", "y"}
_FALSE_WORDS = {"خیر", "نه", "false", "no", "0", "n", ""}
_ACTIVE_WORDS = {"فعال", "active", "1", ""}
_INACTIVE_WORDS = {"غیرفعال", "غیر فعال", "inactive", "0"}

_USERNAME_RE = re.compile(r"^[a-zA-Z0-9._-]{3,50}$")
# سال شمسی تا ۱۵۰۰ و میلادی از ۱۹۰۰ به بعد — برای تشخیص تقویم از روی خود عدد
_JALALI_MAX_YEAR = 1500


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        # نیم‌فاصله و فاصله‌های نامرئی اکسل، مقایسهٔ رشته‌ای را بی‌صدا خراب می‌کنند
        return value.replace("‌", "‌").replace("\xa0", " ").strip()
    return str(value).strip()


def _normalise_digits(value: str) -> str:
    return value.translate(_DIGITS)


def parse_flexible_date(raw: object) -> date | None:
    """تاریخ شمسی یا میلادی، با ارقام فارسی یا اسکی، یا تاریخِ خودِ اکسل.

    سال تشخیص تقویم را می‌دهد: چیزی حوالی ۱۴۰۵ شمسی است و حوالی ۲۰۲۶ میلادی.
    این ابهام واقعی ندارد چون بازه‌ها اصلاً هم‌پوشانی ندارند.
    """
    if raw is None or _text(raw) == "":
        return None
    # اکسل ممکن است سلول را از قبل به datetime تبدیل کرده باشد
    if isinstance(raw, date):
        return raw
    if hasattr(raw, "date"):
        return raw.date()

    cleaned = _normalise_digits(_text(raw)).replace("-", "/").replace(".", "/")
    parts = [p for p in cleaned.split("/") if p]
    if len(parts) != 3:
        return None
    try:
        year, month, day = (int(p) for p in parts)
    except ValueError:
        return None

    try:
        if year <= _JALALI_MAX_YEAR:
            return jdatetime.date(year, month, day).togregorian()
        return date(year, month, day)
    except ValueError:
        return None


def _parse_bool(raw: object) -> bool | None:
    value = _normalise_digits(_text(raw)).lower()
    if value in _TRUE_WORDS:
        return True
    if value in _FALSE_WORDS:
        return False
    return None


def _parse_status(raw: object) -> PersonnelStatus | None:
    value = _normalise_digits(_text(raw)).lower()
    if value in _ACTIVE_WORDS:
        return PersonnelStatus.active
    if value in _INACTIVE_WORDS:
        return PersonnelStatus.inactive
    return None


@dataclass
class ImportRow:
    """یک ردیف فایل، همراه با هرچه اشکال دارد.

    `row_number` شمارهٔ ردیف در خود اکسل است (با احتساب سرستون) تا کاربر بتواند
    مستقیم برود همان‌جا؛ «ردیف سوم دادهٔ معتبر» به درد کسی نمی‌خورد.
    """

    row_number: int
    personnel_code: str = ""
    full_name: str = ""
    job_title: str = ""
    org_unit: str = ""
    is_manager: bool = False
    status: PersonnelStatus = PersonnelStatus.active
    contract_start_date: date | None = None
    contract_end_date: date | None = None
    username: str | None = None
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


@dataclass
class ImportPreview:
    rows: list[ImportRow]
    """خطاهای مربوط به کل فایل (ستون جاافتاده، فایل خالی) — نه یک ردیف خاص."""
    file_errors: list[str] = field(default_factory=list)

    @property
    def valid(self) -> list[ImportRow]:
        return [r for r in self.rows if r.ok]

    @property
    def invalid(self) -> list[ImportRow]:
        return [r for r in self.rows if not r.ok]


def _read_sheet(content: bytes) -> tuple[list[str], list[tuple[int, tuple]]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return [], []
    header = [_text(c) for c in rows[0]]
    body = [(i + 2, r) for i, r in enumerate(rows[1:])]
    return header, body


def parse_workbook(content: bytes, db: Session) -> ImportPreview:
    """فایل را می‌خواند، هر ردیف را اعتبارسنجی می‌کند و گزارش می‌دهد. چیزی نمی‌نویسد."""
    try:
        header, body = _read_sheet(content)
    except Exception:  # noqa: BLE001 — هر خرابی فایل، یک پیام کاربرپسند می‌شود
        return ImportPreview(rows=[], file_errors=["فایل خوانده نشد؛ باید یک فایل معتبر Excel (.xlsx) باشد."])

    if not header:
        return ImportPreview(rows=[], file_errors=["فایل خالی است."])

    index = {name: i for i, name in enumerate(header)}
    missing = [c for c in REQUIRED_COLUMNS if c not in index]
    if missing:
        return ImportPreview(
            rows=[],
            file_errors=[f"این ستون‌ها در فایل نیستند: {'، '.join(missing)}"],
        )

    def cell(row: tuple, column: str) -> object:
        position = index.get(column)
        if position is None or position >= len(row):
            return None
        return row[position]

    # یک‌بار خوانده می‌شوند تا برای هر ردیف یک کوئری جدا نزنیم
    existing_codes = {c for (c,) in db.execute(select(Personnel.personnel_code))}
    existing_usernames = {u for (u,) in db.execute(select(User.username))}

    seen_codes: dict[str, int] = {}
    seen_usernames: dict[str, int] = {}
    parsed: list[ImportRow] = []

    for number, raw in body:
        # ردیف کاملاً خالی (دنبالهٔ سطرهای خالی انتهای فایل) نه خطاست نه داده
        if all(_text(c) == "" for c in raw):
            continue

        item = ImportRow(row_number=number)
        item.personnel_code = _normalise_digits(_text(cell(raw, "کد پرسنلی")))
        item.full_name = _text(cell(raw, "نام و نام خانوادگی"))
        item.job_title = _text(cell(raw, "عنوان شغلی"))
        item.org_unit = _text(cell(raw, "واحد سازمانی"))

        for label, value in (
            ("کد پرسنلی", item.personnel_code),
            ("نام و نام خانوادگی", item.full_name),
            ("عنوان شغلی", item.job_title),
            ("واحد سازمانی", item.org_unit),
        ):
            if not value:
                item.errors.append(f"«{label}» خالی است")

        if item.personnel_code:
            if item.personnel_code in existing_codes:
                item.errors.append("کد پرسنلی از قبل در سامانه ثبت شده است")
            elif item.personnel_code in seen_codes:
                item.errors.append(
                    f"کد پرسنلی تکراری است (ردیف {seen_codes[item.personnel_code]} همین فایل)"
                )
            else:
                seen_codes[item.personnel_code] = number

        is_manager = _parse_bool(cell(raw, "مدیر"))
        if is_manager is None:
            item.errors.append("ستون «مدیر» باید «بله» یا «خیر» باشد")
        else:
            item.is_manager = is_manager

        status = _parse_status(cell(raw, "وضعیت"))
        if status is None:
            item.errors.append("ستون «وضعیت» باید «فعال» یا «غیرفعال» باشد")
        else:
            item.status = status

        item.contract_start_date = parse_flexible_date(cell(raw, "شروع قرارداد"))
        item.contract_end_date = parse_flexible_date(cell(raw, "پایان قرارداد"))
        if item.contract_start_date is None:
            item.errors.append("«شروع قرارداد» خوانده نشد (نمونهٔ درست: ۱۴۰۵/۰۱/۰۱)")
        if item.contract_end_date is None:
            item.errors.append("«پایان قرارداد» خوانده نشد (نمونهٔ درست: ۱۴۰۶/۰۱/۰۱)")
        if (
            item.contract_start_date
            and item.contract_end_date
            and item.contract_end_date <= item.contract_start_date
        ):
            item.errors.append("«پایان قرارداد» باید بعد از «شروع قرارداد» باشد")

        username = _text(cell(raw, "نام کاربری"))
        if username:
            if not _USERNAME_RE.match(username):
                item.errors.append(
                    "نام کاربری فقط می‌تواند حروف انگلیسی، رقم، نقطه، خط تیره و زیرخط داشته باشد (۳ تا ۵۰ نویسه)"
                )
            elif username in existing_usernames:
                item.errors.append("این نام کاربری از قبل وجود دارد")
            elif username in seen_usernames:
                item.errors.append(
                    f"نام کاربری تکراری است (ردیف {seen_usernames[username]} همین فایل)"
                )
            else:
                seen_usernames[username] = number
                item.username = username

        parsed.append(item)

    if not parsed:
        return ImportPreview(rows=[], file_errors=["فایل هیچ ردیف داده‌ای ندارد."])
    return ImportPreview(rows=parsed)


def build_template() -> bytes:
    """فایل نمونهٔ خالی با همان ستون‌ها و یک ردیف راهنما.

    بدون این، کاربر باید ستون‌ها را حدس بزند و اولین تلاشش تقریباً همیشه به خطای
    «ستون جاافتاده» می‌خورد.
    """
    from app.services.excel import _new_sheet, _to_bytes

    workbook, sheet = _new_sheet("پرسنل", COLUMNS)
    sheet.append(
        [
            # عمداً از بازهٔ کدهای دموی seed (P-1001 تا P-1003) فاصله دارد: وقتی
            # نمونه با یک پرسنل واقعی هم‌کد شود، همان فایلی که قرار است راهنما
            # باشد با خطای «کد پرسنلی تکراری است» رد می‌شود.
            "P-0000",
            "نام نمونه",
            "کارشناس",
            "واحد نمونه",
            "خیر",
            "فعال",
            "۱۴۰۵/۰۱/۰۱",
            "۱۴۰۶/۰۱/۰۱",
            "namuneh",
        ]
    )
    return _to_bytes(workbook)
