"""خروجی Excel از فهرست‌های سامانه (ارزیابی‌ها، پرسنل، برنامه‌های بهبود) برای
گزارش‌گیری منابع انسانی."""
from datetime import date
from io import BytesIO

import jdatetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.models.audit_log import AuditLog
from app.models.evaluation import EvaluationRecord
from app.models.improvement_plan import ImprovementPlan
from app.models.personnel import Personnel
from app.models.user import User
from app.services.pdf import to_jalali

_PERSIAN_DIGITS = str.maketrans("0123456789", "۰۱۲۳۴۵۶۷۸۹")


def _jalali_date(value: date | None) -> str:
    """تاریخ (بدون زمان) میلادی → شمسی با ارقام فارسی؛ برای ستون‌های تاریخِ خالص
    مثل قرارداد/بازنگری که «ساعت ۰۰:۰۰» برایشان بی‌معناست."""
    if value is None:
        return ""
    jalali = jdatetime.date.fromgregorian(date=value)
    return jalali.strftime("%Y/%m/%d").translate(_PERSIAN_DIGITS)

_HEADERS = [
    "کد ارزیابی",
    "نام پرسنل",
    "واحد",
    "وضعیت",
    "امتیاز عمومی ٪",
    "امتیاز تخصصی ٪",
    "امتیاز نهایی ٪",
    "نتیجه پیشنهادی",
    "تاریخ شروع",
    "تاریخ نهایی‌شدن",
]

_STATUS_LABELS = {
    "draft": "پیش‌نویس",
    "submitted": "ثبت‌شده",
    "hr_approved": "تأییدشده توسط HR",
    "deputy_approved": "تأییدشده توسط معاونت",
    "finalized": "نهایی‌شده",
}


def _new_sheet(title: str, headers: list[str]) -> tuple[Workbook, "Workbook.worksheets"]:
    """کاربرگ RTL با ردیف سرستون پررنگ و پهنای ستون متناسب — الگوی مشترک همه خروجی‌ها."""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for column_index, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(column_index)].width = max(14, len(header) + 6)
    return wb, ws


def _to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_evaluations_workbook(records: list[EvaluationRecord]) -> bytes:
    wb, ws = _new_sheet("ارزیابی‌ها", _HEADERS)
    for r in records:
        ws.append(
            [
                r.evaluation_code,
                r.subject.full_name,
                r.subject.org_unit,
                _STATUS_LABELS.get(r.status.value, r.status.value),
                float(r.general_score_pct) if r.general_score_pct is not None else None,
                float(r.specialized_score_pct) if r.specialized_score_pct is not None else None,
                float(r.final_weighted_pct) if r.final_weighted_pct is not None else None,
                r.recommendation or "",
                to_jalali(r.created_at.isoformat()),
                to_jalali(r.finalized_at.isoformat()) if r.finalized_at else "",
            ]
        )
    return _to_bytes(wb)


_PERSONNEL_HEADERS = [
    "کد پرسنلی",
    "نام و نام خانوادگی",
    "عنوان شغلی",
    "واحد سازمانی",
    "مدیر",
    "وضعیت",
    "شروع قرارداد",
    "پایان قرارداد",
]


def build_personnel_workbook(rows: list[Personnel]) -> bytes:
    wb, ws = _new_sheet("پرسنل", _PERSONNEL_HEADERS)
    for p in rows:
        ws.append(
            [
                p.personnel_code,
                p.full_name,
                p.job_title,
                p.org_unit,
                "بله" if p.is_manager else "خیر",
                "فعال" if p.status.value == "active" else "غیرفعال",
                _jalali_date(p.contract_start_date),
                _jalali_date(p.contract_end_date),
            ]
        )
    return _to_bytes(wb)


_PLAN_HEADERS = [
    "عنوان برنامه",
    "پرسنل",
    "کد ارزیابی",
    "وضعیت",
    "تاریخ بازنگری",
    "اهداف (انجام‌شده/کل)",
    "مسئول پیگیری",
    "تاریخ ایجاد",
]

_PLAN_STATUS_LABELS = {"open": "باز", "completed": "تکمیل‌شده", "cancelled": "لغوشده"}


_ROLE_LABELS = {
    "unit_supervisor": "مسئول واحد",
    "hr": "منابع انسانی",
    "deputy": "معاونت",
    "ceo": "مدیرعامل",
    "employee": "کارمند",
}

_USER_HEADERS = [
    "نام کاربری",
    "نقش",
    "وضعیت",
    "پرسنل مرتبط",
    "تاریخ ایجاد",
]


def build_users_workbook(users: list[User], personnel_names: dict[int, str]) -> bytes:
    """نام پرسنل مرتبط از دیکشنری ازپیش‌واکشی‌شده می‌آید (کوئری دسته‌ای در روتر از
    N+1 جلوگیری می‌کند)."""
    wb, ws = _new_sheet("کاربران", _USER_HEADERS)
    for u in users:
        ws.append(
            [
                u.username,
                _ROLE_LABELS.get(u.role.value, u.role.value),
                "فعال" if u.is_active else "غیرفعال",
                personnel_names.get(u.personnel_id, "") if u.personnel_id else "",
                to_jalali(u.created_at.isoformat()),
            ]
        )
    return _to_bytes(wb)


_AUDIT_HEADERS = [
    "زمان",
    "کاربر",
    "رویداد",
    "کد ارزیابی",
    "مقدار پیشین",
    "مقدار جدید",
]


def build_audit_log_workbook(
    rows: list[tuple["AuditLog", str | None, str | None]],
    event_labels: dict[str, str],
) -> bytes:
    """هر ردیف: (AuditLog, نام کاربر, کد ارزیابی) — همان JOIN روتر تا از N+1 پرهیز شود."""
    wb, ws = _new_sheet("گزارش رویدادها", _AUDIT_HEADERS)
    for entry, username, evaluation_code in rows:
        ws.append(
            [
                to_jalali(entry.created_at.isoformat()),
                username or f"#{entry.actor_user_id}",
                event_labels.get(entry.event_type, entry.event_type),
                evaluation_code or "",
                str(entry.old_value) if entry.old_value else "",
                str(entry.new_value) if entry.new_value else "",
            ]
        )
    return _to_bytes(wb)


def build_report_workbook(
    *,
    total: int,
    avg_final_pct: float | None,
    by_org_unit: list[tuple[str, float, int]],
    by_indicator: list[tuple[str, str, float, int]],
) -> bytes:
    """گزارش ترکیبی HR در یک فایل با سه برگه: خلاصه، به‌تفکیک واحد، به‌تفکیک شاخص."""
    wb, summary_ws = _new_sheet("خلاصه", ["شاخص", "مقدار"])
    summary_ws.append(["تعداد ارزیابی‌های نهایی‌شده (فیلترشده)", total])
    summary_ws.append([
        "میانگین امتیاز نهایی (٪)",
        float(avg_final_pct) if avg_final_pct is not None else "—",
    ])

    unit_ws = wb.create_sheet("میانگین به‌تفکیک واحد")
    unit_ws.sheet_view.rightToLeft = True
    unit_headers = ["واحد سازمانی", "میانگین امتیاز نهایی ٪", "تعداد"]
    unit_ws.append(unit_headers)
    for cell in unit_ws[1]:
        cell.font = Font(bold=True)
    for column_index, header in enumerate(unit_headers, start=1):
        unit_ws.column_dimensions[get_column_letter(column_index)].width = max(16, len(header) + 6)
    for org_unit, avg, count in by_org_unit:
        unit_ws.append([org_unit, float(avg), count])

    indicator_ws = wb.create_sheet("میانگین به‌تفکیک شاخص")
    indicator_ws.sheet_view.rightToLeft = True
    indicator_headers = ["دسته", "شرح شاخص", "میانگین امتیاز (از ۵)", "تعداد"]
    indicator_ws.append(indicator_headers)
    for cell in indicator_ws[1]:
        cell.font = Font(bold=True)
    for column_index, header in enumerate(indicator_headers, start=1):
        indicator_ws.column_dimensions[get_column_letter(column_index)].width = max(
            18, len(header) + 6
        )
    for category, description, avg, count in by_indicator:
        indicator_ws.append([category, description, float(avg), count])

    return _to_bytes(wb)


def build_improvement_plans_workbook(
    plans: list[ImprovementPlan],
    evaluation_codes: dict[int, str],
    owner_usernames: dict[int, str],
) -> bytes:
    """کد ارزیابی و نام مسئول از دیکشنری‌های ازپیش‌واکشی‌شده می‌آیند (مدل این دو
    relationship را ندارد و کوئری دسته‌ای در روتر از N+1 جلوگیری می‌کند)."""
    wb, ws = _new_sheet("برنامه‌های بهبود", _PLAN_HEADERS)
    for plan in plans:
        done = sum(1 for g in plan.goals if g.is_done)
        ws.append(
            [
                plan.title,
                plan.personnel.full_name,
                evaluation_codes.get(plan.evaluation_record_id, ""),
                _PLAN_STATUS_LABELS.get(plan.status.value, plan.status.value),
                _jalali_date(plan.review_date),
                f"{done}/{len(plan.goals)}",
                owner_usernames.get(plan.owner_user_id, "") if plan.owner_user_id else "",
                to_jalali(plan.created_at.isoformat()),
            ]
        )
    return _to_bytes(wb)
